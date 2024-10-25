# Import openpyxl
from openpyxl import Workbook

# Read the file
def main():
    # Open the text file
    with open("C:\\Users\\DELL\\OneDrive\\Desktop\\Flight_Details_make_my_trip.txt", 'r', encoding="utf-8") as file:
        lines = file.readlines()

    # Split the file content into individual datasets using the separator
    datasets = []
    currentDataset = []

    for line in lines:
        if "--------------------------------------------------------" in line:
            if currentDataset:
                datasets.append(currentDataset)
                currentDataset = []
        else:
            currentDataset.append(line)

    # Don't forget to add the last dataset
    if currentDataset:
        datasets.append(currentDataset)

    # Create a new Excel workbook and sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Flight Details"

    # Create header row
    header = ["Date", "Flight Name", "Flight ID", "Price", "Departure Airport", "Arrival Airport", "Total Time", "Non-stop/With Stops", "Departure Time", "Arrival Time"]
    sheet.append(header)

    # Variable to keep track of row number in the Excel file
    rowNum = 1

    # Iterate over each dataset and extract the details
    for dataset in datasets:
        date = "Unknown"
        departureTime = "Unknown"
        arrivalTime = "Unknown"
        fromLocation = "Unknown"
        toLocation = "Unknown"
        flightCarrier = "Unknown"
        ticketPrice = "Unknown"
        flightDuration = "Unknown"
        flightId = "Unknown"
        nonStopOrWithStops = "Unknown"

        for index, line in enumerate(dataset):
            # Date extraction
            if index == 0 and "Date:" in line:
                date = line.split(":")[1].strip()

            # Flight carrier extraction
            if index == 1 and "Flight Name:" in line:
                flightCarrier = line.split(":")[1].strip()

            # Flight ID extraction
            if index == 2 and "Flight ID:" in line:
                flightId = line.split(":")[1].strip()

            # Ticket price extraction
            if index == 3 and "Price:" in line:
                ticketPrice = line.split(":")[1].strip()
                # Remove any currency symbols or commas
                ticketPrice = ticketPrice.replace("₹", "").replace(",", "").strip()

            # Departure Airport extraction
            if index == 4 and "Departure Airport:" in line:
                fromLocation = line.split(":")[1].strip()

            # Arrival Airport extraction
            if index == 5 and "Arrival Airport:" in line:
                toLocation = line.split(":")[1].strip()

            # Total time (flight duration) extraction
            if index == 6 and "Total Time:" in line:
                flightDuration = line.split(":")[1].strip()

            # Non-stop or with stops extraction
            if index == 7:
                nonStopOrWithStops = line.strip()

            # Departure Time extraction
            if index == 8 and "Departure Time:" in line:
                departureTime = line.split("Departure Time:")[1].strip()

            # Arrival Time extraction
            if index == 9 and "Arrival Time:" in line:
                arrivalTime = line.split("Arrival Time:")[1].strip()

        # Write the extracted data to the Excel file
        row = [date, flightCarrier, flightId, ticketPrice, fromLocation, toLocation, flightDuration, nonStopOrWithStops, departureTime, arrivalTime]
        sheet.append(row)

        # Convert ticket price to numeric if it's a valid number
        if ticketPrice.replace('.', '', 1).isdigit():
            sheet.cell(row=rowNum + 1, column=4).value = float(ticketPrice)  # Save as numeric value
        else:
            sheet.cell(row=rowNum + 1, column=4).value = ticketPrice  # Save as text if not numeric

        rowNum += 1

    # Save the Excel file
    workbook.save("C:\\Users\\DELL\\OneDrive\\Desktop\\Flight_Details_make_my_trip_data.xlsx")

    print("All flight details saved to Excel with proper ticket price formatting!")

if __name__ == "__main__":
    main()
